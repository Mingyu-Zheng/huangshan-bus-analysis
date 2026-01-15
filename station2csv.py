#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Line XLSX files -> CSV converter.

Each workbook under ``线路/线路`` describes one bus line where every record is a
stop encoded with the provider's proprietary coordinate scheme.  This script
walks every workbook, decodes the coordinates, normalizes metadata and finally
writes a CSV that matches the simulator input format.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, Iterator, List, Sequence, Tuple

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent
LINES_DIR = REPO_ROOT / "INIT_station_excel"
OUTPUT_DIR = REPO_ROOT / "MID_output"
DEFAULT_OUTPUT = OUTPUT_DIR / "huangshan.csv"

Stop = Tuple[str, int, str, str, float, float]

LAT_OFFSET = 11_600_000
LON_OFFSET = 47_200_000


def _decode_coord(value: object, offset: int) -> float | None:
    """
    Reverse ``INT(coord)*1e6 + (coord - INT(coord)) * 0.6 * 1e6`` after removing
    the fixed offsets used by the data provider.
    """

    if value is None or value == "":
        return None
    try:
        scaled = int(float(value))
    except (TypeError, ValueError):
        return None

    scaled += offset
    degrees = scaled // 1_000_000
    remainder = scaled - degrees * 1_000_000
    if remainder >= 600_000:
        return None

    fractional = remainder / 600_000.0
    return round(degrees + fractional, 8)


def _normalize_line_id(raw: object, fallback: str) -> str:
    if raw is None or str(raw).strip() == "":
        raw_str = fallback
    else:
        raw_str = str(raw).strip()
    try:
        return f"{int(float(raw_str)):04d}"
    except ValueError:
        return raw_str


def _normalize_direction(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if "主" in text:
        return "A"
    if "副" in text:
        return "B"
    return None


def iter_line_workbooks(lines_dir: Path = LINES_DIR) -> Iterator[Path]:
    if not lines_dir.exists():
        return
    for path in sorted(lines_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        yield path


def parse_workbook(path: Path) -> List[Stop]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
        header_map = {name: idx for idx, name in enumerate(headers)}
        required = ["线路", "站牌号", "站点名称", "发车方向", "经度", "纬度"]
        if not all(key in header_map for key in required):
            return []

        stops: List[Stop] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[header_map["站点名称"]]
            if not name_raw:
                continue
            station_name = str(name_raw).strip()
            if not station_name:
                continue

            line_id = _normalize_line_id(row[header_map["线路"]], path.stem)
            station_idx = row[header_map["站牌号"]]
            try:
                station_id = int(float(station_idx))
            except (TypeError, ValueError):
                continue

            direction = _normalize_direction(row[header_map["发车方向"]])
            if not direction:
                continue

            lon = _decode_coord(row[header_map["经度"]], LON_OFFSET)
            lat = _decode_coord(row[header_map["纬度"]], LAT_OFFSET)
            if lon is None or lat is None:
                continue

            stops.append((line_id, station_id, direction, station_name, lon, lat))
        return stops
    finally:
        wb.close()


def collect_stops(workbooks: Iterable[Path]) -> List[Stop]:
    collected: List[Stop] = []
    for workbook in workbooks:
        collected.extend(parse_workbook(workbook))
    # Sort to keep deterministic output: by line, direction, station id
    collected.sort(key=lambda item: (item[0], item[2], item[1], item[3]))
    return collected


def write_csv(stops: Sequence[Stop], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def fmt(value: float) -> str:
        return f"{value:.8f}".rstrip("0").rstrip(".")

    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        csv_file.write(
            "line_id,station_id,direction,station_name,station_lon,station_lat\n"
        )
        for line_id, station_id, direction, name, lon, lat in stops:
            csv_file.write(
                f"{line_id},{station_id},{direction},{name},{fmt(lon)},{fmt(lat)}\n"
            )


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Convert line XLSX files to CSV.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Destination CSV path (default: %(default)s)",
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=LINES_DIR,
        help="Directory containing XLSX line files (default: %(default)s)",
    )
    args = parser.parse_args(argv)

    stops = collect_stops(iter_line_workbooks(args.input))
    write_csv(stops, args.output)


if __name__ == "__main__":
    main()
